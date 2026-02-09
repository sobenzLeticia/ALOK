import streamlit as st
import pandas as pd
import datetime as dt
from collections import defaultdict
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter


import base64


@st.cache_data
def load_salas(uploaded_file):
    return pd.read_excel(uploaded_file)

@st.cache_data
def load_turmas(uploaded_file):
    return pd.read_excel(uploaded_file)

# ================= INTERFACE =================
st.set_page_config(page_title="Alocação de Turmas", layout="wide")
st.title("ALOK - Sistema de Apoio à Decisão de Alocação de Turmas em Salas")

file_salas = st.file_uploader("📂 Envie o arquivo de SALAS", type=["xlsx"])
file_turmas = st.file_uploader("📂 Envie o arquivo de TURMAS", type=["xlsx"])

# Inicializar variáveis de sessão
if "resultados" not in st.session_state:
    st.session_state.resultados = None
if "buffers_salas" not in st.session_state:
    st.session_state.buffers_salas = {}

# ================= PROCESSAMENTO =================
if file_salas and file_turmas:
    if st.button("🚀 Rodar Alocação"):
        # ================= LEITURA =================
        df_salas = load_salas(file_salas)
        df_turmas = load_turmas(file_turmas)

        salas_data = df_salas["SALAS"].to_numpy()
        capacidade_data = df_salas["CAPACIDADE"].to_numpy()
        acesso_sala = df_salas["ACESSIBILIDADE"].to_numpy()
        cod_data = df_turmas["CÓDIGO"].to_numpy()
        cod_turma_data = df_turmas["Nº DA TURMA"].to_numpy()
        turmas_data = df_turmas["DISCIPLINA"].to_numpy()
        demanda_data = df_turmas["PREVISÃO DE ALUNOS"].to_numpy()
        professor_data = df_turmas["PROFESSOR"].to_numpy()
        dias_data = df_turmas["DIAS"].to_numpy()
        horarios_data = df_turmas["HORÁRIOS"].to_numpy()
        horario_ini = df_turmas["HÓRARIO INÍCIO"].to_numpy()
        horario_fim = df_turmas["HÓRARIO FINAL"].to_numpy()
        curso = df_turmas["CURSO"].to_numpy()
        acessibilidade = df_turmas["ACESSIBILIDADE"].to_numpy()

        horario_inicio = []
        horario_final = []

        for i in range(len(horarios_data)):
            horario_inicio.append(horario_ini[i])
            horario_final.append(horario_fim[i])


        # Montar horários por turma
        horarios_turmas = []
        for i in range(len(dias_data)):
            dias_list = str(dias_data[i]).split()
            horarios_list = str(horarios_data[i]).split(', ')
            turma_horarios = []
            for dia in dias_list:
                for hora in horarios_list:
                    turma_horarios.append(f'{dia} {hora}')
            horarios_turmas.append(turma_horarios)

        # Criar lista de salas
        salas_ct = []
        for i in range(len(salas_data)):
            salas_ct.append({
                "NOME": salas_data[i],
                "CAPACIDADE": capacidade_data[i],
                "HORARIOS_OCUPADOS": set(),
                "ACESSEBILIDADE_SALA": acesso_sala[i]
            })

        # Criar lista de disciplinas
        disciplinas = []
        for i in range(len(turmas_data)):
            disciplinas.append({
                "CURSO": curso[i],
                "CODIGO": cod_data[i],
                "DISCIPLINA": turmas_data[i],
                "CODIGO TURMA": cod_turma_data[i],
                "DIAS": dias_data[i],
                "HORARIO INICIO": horario_inicio[i],
                "HORARIO FINAL": horario_final[i],
                "HORARIOS": horarios_turmas[i],
                "ALUNOS": demanda_data[i],
                "PROFESSOR": professor_data[i],
                "ACESSIBILIDADE": acessibilidade[i]
            })

        disciplinas.sort(key=lambda d: d["ALUNOS"], reverse=True)
       
        sala_prof = []

        # ================= ALOCAÇÃO =================
        alocacao = []
        for disc in disciplinas:
            alunos = disc["ALUNOS"]
            horarios_disciplina = disc["HORARIOS"]
            
            melhor_sala = None
            menor_ociosidade = float("inf")
            
            professor = disc["PROFESSOR"]
            dias_usados = disc["DIAS"]
            horario_seguinte_inicio = disc["HORARIO INICIO"]

            acessibilidade = disc["ACESSIBILIDADE"]

            if professor in sala_prof: 
                if horario_seguinte_inicio in professor["HORARIOS FINAL"] and dias_usados == professor["DIAS"]:
                    sala_usada = professor["SALA"]
                
                    for sala in salas_ct:
                        is_disponivel = all(h not in sala["HORARIOS_OCUPADOS"] for h in horarios_disciplina)
                        if sala["NOME"] == sala_usada and is_disponivel:
                            ociosidade = sala["CAPACIDADE"] - alunos
                            if ociosidade < menor_ociosidade:
                                menor_ociosidade = ociosidade
                                melhor_sala = sala

            if acessibilidade in "SsIiMm" and not melhor_sala:
                for sala in salas_ct:    
                    is_acessivel = True if sala["ACESSIBILIDADE"] in "SsIiMm" else False
                    is_disponivel = all(h not in sala["HORARIOS_OCUPADOS"] for h in horarios_disciplina)
                    if is_acessivel and is_disponivel:
                        ociosidade = sala["CAPACIDADE"] - alunos
                        melhor_sala = sala

            if not melhor_sala:
                for sala in salas_ct:
                    if sala["CAPACIDADE"] >= alunos:
                        is_disponivel = all(h not in sala["HORARIOS_OCUPADOS"] for h in horarios_disciplina)
                        if is_disponivel:
                            ociosidade = sala["CAPACIDADE"] - alunos
                            if ociosidade < menor_ociosidade:
                                menor_ociosidade = ociosidade
                                melhor_sala = sala

            if melhor_sala:
                alocacao.append({
                    "CURSO": disc["CURSO"],
                    "CODIGO": disc["CODIGO"],
                    "DISCIPLINA": disc["DISCIPLINA"],
                    "SALA": melhor_sala["NOME"],
                    "TURMA": disc["CODIGO TURMA"],
                    "PROFESSOR": disc["PROFESSOR"],
                    "DIAS": disc["DIAS"],
                    "HORARIO INICIO": disc["HORARIO INICIO"],
                    "HORARIO FINAL": disc["HORARIO FINAL"],
                    "HORARIO": ", ".join(horarios_disciplina),
                    "ALUNOS": alunos,
                    "OCIOSIDADE": menor_ociosidade,
                    "STATUS": "Alocada"
                })
                for h in horarios_disciplina:
                    melhor_sala["HORARIOS_OCUPADOS"].add(h)
                sala_prof.append({
                    "PROFESSOR": disc["PROFESSOR"],
                    "SALA": melhor_sala["NOME"],
                    "DIAS": disc["DIAS"],
                    "HORARIO FINAL": disc["HORARIO FINAL"]
                })
                
            else:
                alocacao.append({
                    "CURSO": disc["CURSO"],
                    "CODIGO": disc["CODIGO"],
                    "DISCIPLINA": disc["DISCIPLINA"],
                    "SALA": None,
                    "TURMA": disc["CODIGO TURMA"],
                    "PROFESSOR": disc["PROFESSOR"],
                    "DIAS": disc["DIAS"],
                    "HORARIO INICIO": disc["HORARIO INICIO"],
                    "HORARIO FINAL": disc["HORARIO FINAL"],
                    "HORARIO": ", ".join(horarios_disciplina),
                    "ALUNOS": alunos,
                    "OCIOSIDADE": menor_ociosidade,
                    "STATUS": "Não alocada"
                })


        # Criar e salvar o DataFrame de resultados
        df_resultados = pd.DataFrame(alocacao)
        buffer_geral = BytesIO()
        df_resultados.to_excel(buffer_geral, index=False)
        buffer_geral.seek(0)

        st.session_state.resultados = buffer_geral

        dias_semana = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado"]
        horas_minutos = []
        for h in range(7, 22):
            horas_minutos.append(f'{h:02d}:00 - {h:02d}:30')
            horas_minutos.append(f'{h:02d}:30 - {h+1:02d}:00')

        def split_horario(horario_completo):
            partes = horario_completo.split()
            dia = partes[0]
            hora_str = partes[1]
            if '-' not in hora_str:
                return []
            hora_inicio_str, hora_fim_str = hora_str.split('-')
            hora_inicio = dt.datetime.strptime(hora_inicio_str, '%H:%M:%S')
            intervalos = []
            intervalos.append(f'{dia} {hora_inicio.strftime("%H:%M")} - {hora_inicio.replace(minute=30).strftime("%H:%M")}')
            hora_segundo_intervalo = hora_inicio.replace(minute=30)
            intervalos.append(f'{dia} {hora_segundo_intervalo.strftime("%H:%M")} - {dt.datetime.strptime(hora_fim_str, "%H:%M:%S").strftime("%H:%M")}')
            return intervalos

        horarios_por_sala = defaultdict(lambda: defaultdict(dict))
        for aloc in alocacao:
            if aloc['SALA']:
                sala_nome = aloc['SALA']
                disciplina_info = f"{aloc['CODIGO']} - {aloc['DISCIPLINA']} - {aloc['TURMA']} - {aloc['PROFESSOR']}"
                horarios_blocos = [h.strip() for h in aloc['HORARIO'].split(',')]
                for bloco in horarios_blocos:
                    if bloco:
                        dia = bloco.split()[0]
                        horarios_30min = split_horario(bloco)
                        for horario_30min in horarios_30min:
                            _, horario_formatado = horario_30min.split(' ', 1)
                            horarios_por_sala[sala_nome][dia][horario_formatado] = disciplina_info

        borda_fina = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
        alinhamento_centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
        fonte_padrao = Font(size=10)

        wb = Workbook()
        wb.remove(wb.active)

        for sala in salas_ct:
            sala_nome = sala["NOME"]
            ws = wb.create_sheet(title=sala_nome[:31])
            ws.title = "Horário"+ sala_nome[:25]

            # Colocar o CT | SALA | CAPACIDADE
            info_sala = f"Centro de Tecnologia | {sala_nome} | Capacidade: {sala['CAPACIDADE']}"
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(dias_semana) + 1)
            cell_info = ws.cell(row=1, column=1, value=info_sala)
            cell_info.font = Font(bold=True, size=12)
            cell_info.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=2, column=1, value="Horário").font = Font(bold=True)
            for col, dia in enumerate(dias_semana, start=2):
                ws.cell(row=2, column=col, value=dia).font = Font(bold=True)

            for row, hora in enumerate(horas_minutos, start=3):
                ws.cell(row=row, column=1, value=hora)

            if sala_nome in horarios_por_sala:
                for dia, horarios in horarios_por_sala[sala_nome].items():
                    if dia == 'SEGUNDA': col = 2
                    elif dia == 'TERÇA': col = 3
                    elif dia == 'QUARTA': col = 4
                    elif dia == 'QUINTA': col = 5
                    elif dia == 'SEXTA': col = 6
                    elif dia == 'SÁBADO': col = 7
                    else: continue
                    for horario, info in horarios.items():
                        if horario in horas_minutos:
                            row_idx = horas_minutos.index(horario) + 3
                            ws.cell(row=row_idx, column=col, value=info)

            # Mesclar células
            for col in range(2, len(dias_semana) + 2):
                start_row = 3
                current_value = ws.cell(row=3, column=col).value
                for row in range(3, len(horas_minutos) + 3):
                    value = ws.cell(row=row, column=col).value
                    if value != current_value:
                        if current_value not in (None, "") and row - 1 > start_row:
                            ws.merge_cells(start_row=start_row, start_column=col,
                                            end_row=row - 1, end_column=col)
                        start_row = row
                        current_value = value
                if current_value not in (None, "") and len(horas_minutos) + 2 > start_row:
                    ws.merge_cells(start_row=start_row, start_column=col,
                                    end_row=len(horas_minutos) + 2, end_column=col)

            # Estilo
            for row in ws.iter_rows(min_row=1, max_row=len(horas_minutos) + 2,
                                    min_col=1, max_col=len(dias_semana) + 1):
                for cell in row:
                    cell.border = borda_fina
                    cell.alignment = alinhamento_centro
                    cell.font = fonte_padrao

            for col in range(1, len(dias_semana) + 2):  # +1 porque tem a coluna de horários
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 20  
  

        # Salvar
        buffer_salas = BytesIO()
        wb.save(buffer_salas)
        buffer_salas.seek(0)

        # Guarda no session_state como 1 único arquivo
        st.session_state.buffers_salas = {"Horarios_por_sala.xlsx": buffer_salas}


if st.session_state.resultados:
    st.download_button(
        label="⬇️ Baixar Resultados Gerais",
        data=st.session_state.resultados,
        file_name="Resultados_Gerais.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    for nome_arquivo, buffer in st.session_state.buffers_salas.items():
        st.download_button(
            label=f"⬇️ Baixar {nome_arquivo}",
            data=buffer,
            file_name=nome_arquivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

with st.expander("**SOBRE O SISTEMA**", expanded=False):
    st.markdown("""
    <div style="padding: 1rem;">
        <h3> ALOK - Sistema de Apoio à Decisão de Alocação de Turmas</h3>
        
        Funcionalidades:
        - Alocação Inteligente: Distribuição automática de turmas em salas
        - Otimização: Minimiza ociosidade e maximiza utilização
        - Relatórios: Geração de horários e relatórios detalhados
        
        Como usar:        
        1. Faça upload dos arquivos de **Salas** e **Turmas**
        2. Clique em **Executar Alocação Automática**
        3. Baixe os relatórios gerados
        
        Desenvolvimento:
        Autora: Ana Letícia S. B. de Menezes  
        Orientador: Bruno de Athayde Prata  
        
        ---
        Sistema desenvolvido para otimização de recursos educacionais
    </div>
    """, unsafe_allow_html=True)
